import openpyxl
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import statistics as stats

'''
EXCEL
1) Ejercicio de manejo de Excel en base al fichero 02_Excel_data.xlsx:

Abre el fichero Excel
Lista las hojas que hay
Crea una nueva hoja que se llame olimpiadas entre ventas y otros
Vuelve a listar las hojas que hay
Guarda los cambios y comprueba en el excel si se han efectuado
'''

wb = openpyxl.load_workbook("Data manipulation/02_Excel_data.xlsx")
def nombre_hojas():
    print(f"Nombres de hojas: {wb.sheetnames}")
    
nombre_hojas()
hoja_olimpiadas = wb.create_sheet("olimpiadas", 2)
nombre_hojas()
wb.save("Data manipulation/02_Excel_data.xlsx")

'''
2) Crea una lista que contenga los siguientes datos olímpicos (nombre del país y medallas de oro, plata y bronce):
USA, 46, 12, 5
China, 38, 20, 7
UK, 29, 7, 7
Russia, 22, 10, 9
South Korea, 13, 3, 2
Germany, 11, 7, 4
Añadir esa lista a la hoja de olimpiadas que creamos en el ejercicio anterior
Listar por filas y columnas el contenido
Mover todo el rango de datos una fila hacia abajo (buscar en la documentación como mover rangos)
Añadir unas cabeceras en la primera línea (que acaba de quedar libre) que sean "Pais", "Oros", "Platas", "Bronces"
Guarda los cambios y comprueba en el excel si se han efectuado
'''

lista_olimpicos = [["USA", 46, 12, 5], ["China", 38, 20, 7], ["UK", 29, 7, 7], ["Russia", 22, 10, 9], ["South Korea", 13, 3, 2], ["Germany", 11, 7, 4]]
for data in lista_olimpicos:
    hoja_olimpiadas.append(data)

for fila in hoja_olimpiadas.rows:
    for columna in fila:
        if columna.value != None:
            print(columna.coordinate, columna.value)  # Listar por filas y columnas el contenido
    print("----Final de Fila---")

hoja_olimpiadas.move_range("A1:D6", rows=1) # Mover todo el rango de datos una fila hacia abajo

def add_header(col, val):
    hoja_olimpiadas.cell(row=1, column=col, value=val) # Añadir cabeceras en la 1ª línea

add_header(1, "País")
add_header(2, "Oros")
add_header(3, "Platas")
add_header(4, "Bronces")

wb.save("Data manipulation/02_Excel_data.xlsx")

'''
3) Continuemos con Excel, centrandonos en la parte de la librería openpyxl.styles (la cual proporciona estilo a nuestras hojas de Excel)
Poner las cabeceras que acabamos de añadir en negrita (bold)
Guarda los cambios y comprueba en el excel si se han efectuado
'''

font_style = Font(bold=True)
headers = [hoja_olimpiadas['A1'], hoja_olimpiadas['B1'], hoja_olimpiadas['C1'], hoja_olimpiadas['D1']]

def bold_headers():
    for x in headers:
        x.font = font_style

bold_headers()
        
wb.save("Data manipulation/02_Excel_data.xlsx")

'''4) Continuamos con Excel. Crea una nueva columna que sea el sumatorio de todas las medallas conseguidas por cada uno de los paises'''
hoja_olimpiadas['E1'] = "Sumatorio"

def suma():
    for i in range(2,8,1):
        hoja_olimpiadas[f"E{i}"].value = f"=SUM(B{i}:D{i})"

suma()

wb.save("Data manipulation/02_Excel_data.xlsx")

'''5) Continuamos con Excel. Crear un gráfico de barras (openpyxl.chart.BarChart()) lo más parecido al que se muestra en la imagen:'''
wb = openpyxl.load_workbook("Data manipulation/02_Excel_data.xlsx")
hoja_olimpiadas = wb["olimpiadas"]

grafica = openpyxl.chart.BarChart()
grafica.title = "Medallas Olimpicas"
grafica.varyColors = "0000FFFF"

data = openpyxl.chart.Reference(hoja_olimpiadas, min_col=5, min_row=1, max_col=5, max_row=7)
categs = openpyxl.chart.Reference(hoja_olimpiadas, min_col=1, min_row=2, max_row=7)

grafica.add_data(data, titles_from_data=True)
grafica.set_categories(categs)

hoja_olimpiadas.add_chart(grafica, "A9")
wb.save("Data manipulation/02_Excel_data.xlsx")

'''
CSV
Vamos a realizar un ejercicio práctico de minería de datos, donde limpiaremos y filtraremos información con un csv que contiene datos de vehículos
Trabajaremos con el fichero 02_CSV_data.csv

6) Observa los datos del csv y realiza las siguientes tareas:
Tenemos 9 columnas, las 8 primeras contienen datos numéricos con los cuales podemos trabajar, crea una lista para cada una de estas columnas (mpg, cylinders, etc.). No es necesario crear lista para name
Recorre los datos del csv adecuadamente y almacena los datos de cada columna en cada una de las listas que has creado anteriormente.
Comprobar que se haya guardado en las listas la información correspondiente (con prints)
'''

import csv
with open("Data manipulation/02_CSV_data.csv", "r") as f:
    reader = csv.reader(f, delimiter=',')

    mpg = []
    cylinders = []
    displacement = []
    horsepower = []
    weight = []
    acceleration = []
    year = []
    origin = []

    for row in reader:
        mpg.append(row[0])
        cylinders.append(row[1])
        displacement.append(row[2])
        horsepower.append(row[3])
        weight.append(row[4])
        acceleration.append(row[5])
        year.append(row[6])
        origin.append(row[7])

    lista_general = [mpg, cylinders, displacement, horsepower, weight, acceleration, year, origin]
    for lista in lista_general:
        print(f"list of {lista[0]}:", lista, "\n")

'''
7) Vamos a trabajar con los números de nuestras listas por lo que nos sobra el primer elemento de cada lista, el cual contiene el nombre de cada una de las columnas.

Elimina el primer elemento de cada una de las listas
Observa que los datos de nuestras listas son strings, conviértelos a float. Pista: list y map te pueden ayudar.
Comprobar que se haya eliminado el primer elemento y que los datos son numéricos (con prints)
'''

dictionary = {'mpg':mpg,
              'cylinders':cylinders,
              'displacement':displacement,
              'horsepower':horsepower,
              'weight':weight,
              'acceleration':acceleration,
              'year':year,
              'origin':origin}
    
lista_general_float = []
    
for nombre_lista, lista in dictionary.items():
    lista.pop(0)
    lista_float = list(map(float,lista))
    lista_general_float.append(lista_float)
    print(f"list of {nombre_lista}:", lista_float, "\n")

'''
8) Para terminar vamos a mostrar algunos resultados. Para ello utilizaremos la librería statistics.

Muestra para cada una de las listas, cuantas observaciones tiene (cuántos datos tiene).
Mostrar el mínimo y el máximo de cada lista
Mostrar la media, mediana y desviación estándar de cada lista (utilizando statistics)
'''

nombre_lista = []
    
for key, value in dictionary.items():
    print(f"Número de muestras en lista {key}:", len(value))
    nombre_lista.append(key)
    
    # print(nombre_lista)
    
def minimum_maximum():
    contador = 0
    for lista_float in lista_general_float:
        minimum = min(lista_float)
        maximum = max(lista_float)
        print(f"El mínimo de lista {nombre_lista[contador]} es:", minimum)
        print(f"El máximo de lista {nombre_lista[contador]}es", maximum, "\n")
        contador += 1
    
def media():
    contador = 0
    for lista_float in lista_general_float:
        media = stats.mean(lista_float)
        print(f"La media de lista {nombre_lista[contador]} es:", media)
        contador += 1
    
def mediana():
    contador = 0
    for lista_float in lista_general_float:
        mediana = stats.median(lista_float)
        print(f"La mediana de lista {nombre_lista[contador]} es:", mediana)
        contador += 1
    
def desviacion_estandar():
    contador = 0
    for lista_float in lista_general_float:
        desviacion_estandar = stats.stdev(lista_float)
        print(f"La desviación estándar de lista {nombre_lista[contador]} es:", desviacion_estandar)
        contador += 1
            
print("\n-----El mínimo y el máximo-----")
minimum_maximum()
print("\n-----La media-----")
media()
print("\n-----La mediana-----")
mediana()
print("\n-----La desviación estándar-----")
desviacion_estandar()

f.close()